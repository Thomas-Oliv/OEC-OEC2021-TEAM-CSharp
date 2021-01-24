
import os
import random
from statistics import mean
import os.path
from os import path
from openpyxl import Workbook, load_workbook
from Objects.Student import Student
from Objects.Teacher import Teacher
from Objects.TeachingAssistant import TeachingAssistant
from Objects.Directory import Directory


def main():
    # Get project directly
    PROJECT_ROOT = os.path.abspath(os.path.dirname(__file__))
    # Build directory
    filedir = os.path.join(PROJECT_ROOT, 'OEC2021_-_School_Record_Book_.xlsx')

    while True:
        file = input('Enter a file (with directory) to import. Leave blank to use default or \'e\' to exit: ')
        if file == 'e':
            return
        elif not file:
            break
        elif path.exists(file) and file.find('.xlsx') != -1 and file.find('\\') != -1:
            directory = file
            break
        else:
            print("File does not exist!")

    # Empty lists to store students, teachers, and teacher's assistants
    students = list()
    teachers = list()
    tas = list()
    # Get workbook from directory
    wb = load_workbook(read_only=True, filename=filedir)
    # Get the student info sheet from the workbook
    student_info = wb[wb.sheetnames[0]]
    # Iterate through each row in the student workbook
    for student in student_info.iter_rows():
        # If the row is empty
        if not any(student):
            # Break from loop
            break
        # Check for valid student number
        if isinstance(student[0].value, float) and student[0].value > 0:
            # Get classes for the student ordered by period 1-4
            classes = [student[4].value, student[5].value, student[6].value, student[7].value]
            # Check for extracurriculars
            if student[9].value != 'N/A' and student[9].value is not None:
                # Get first extracurriculars from the comma delimited list.
                extra_c = student[9].value.split(',')[0]
            else:
                # If no extracurriculars then set value to None
                extra_c = None
            # Check for health problems and if not none and NA then set value to true
            health_problem = student[8].value is not None and student[8].value != 'N/A'
            # Append student object to students list
            students.append(Student(int(student[0].value), student[2].value, student[1].value, int(student[3].value), health_problem, classes, extra_c))
    # Get teacher sheet
    teacher_info = wb[wb.sheetnames[1]]
    # Iterate through each row in the teacher sheet
    for teacher in teacher_info.iter_rows():
        if not any(teacher):
            # If the row is empty (aka eof reached) then break
            break
        # Check for valid teacher number
        if isinstance(teacher[0].value, float) and teacher[0].value > 0:
            # Append new teacher object to teachers
            teachers.append(Teacher(int(teacher[0].value), teacher[2].value, teacher[1].value, teacher[3].value))
    # Get the teacher assistant sheet
    ta_info = wb[wb.sheetnames[2]]
    # Iterate through the rows in the TA sheet
    for index, ta in enumerate(ta_info.iter_rows()):
        # If the row is empty (aka eof reached) then break
        if not any(ta):
            break
        if index == 0:
            continue
        # Get classes for the TA ordered by period 1-4
        classes = [ta[2].value, ta[3].value, ta[4].value, ta[5].value]
        # Append new TA to the TA list
        tas.append(TeachingAssistant(ta[1].value, ta[0].value, classes))

    # Get infected people sheet
    infected_info = wb[wb.sheetnames[3]]
    # Iterate through all infected people
    for infected in infected_info.iter_rows():
        # If row is empty (aka eof) then break
        if not any(infected):
            break
        # Check for ID. If ID exists then it is a student
        if isinstance(infected[0].value, float) and infected[0].value > 0:
            # Get student from ID (index = student# -1)
            student = students[int(infected[0].value)-1]
            # Validate that the name matches the ID
            if student.first == infected[2].value and student.last == infected[1].value:
                student.chanceOfDisease = 1
            else:
                # Raise exception if they do not match
                raise Exception('Name does not match existing student with specified ID')
        #If Student # is N/A then linearly search TA List and set infected accordingly.
        elif infected[0].value == 'N/A':
            for ta in tas:
                #Match TA from firstname / lastname
                if ta.first == infected[2].value and ta.last == infected[1].value:
                    ta.chanceInfected = 1
                    break

    wb.close()
    directory = Directory(teachers, students, tas)
    directory.reducePeriod(1)
    directory.reducePeriod(2)
    directory.reducePeriod('Lunch')
    directory.reducePeriod(3)
    directory.reducePeriod(4)
    directory.reducePeriod('Extra')

    while True:
        name = input('Type a name or \'e\' to exit. Type \'f\' to output results to excel: ')
        person = None
        if name == 'e':
            break
        elif name == 'f':
            newfile = Workbook()
            ws1 = newfile.create_sheet('Students',0)
            ws1['A1'] = 'Student Number'
            ws1['B1'] = 'Last Name'
            ws1['C1'] = 'First Name'
            ws1['D1'] = 'Infectivity'
            stud_count = 2
            for student in students:
                ws1[f'A{stud_count}'] = student.id
                ws1[f'B{stud_count}'] = student.last
                ws1[f'C{stud_count}'] = student.first
                ws1[f'D{stud_count}'] = student.chanceOfDisease
                stud_count += 1

            ws2 = newfile.create_sheet('Teachers',1)
            ws2['A1'] = 'Teacher Number'
            ws2['B1'] = 'Last Name'
            ws2['C1'] = 'First Name'
            ws2['D1'] = 'Infectivity'

            teach_count = 2
            for teacher in teachers:
                ws2[f'A{teach_count}'] = teacher.id
                ws2[f'B{teach_count}'] = teacher.last
                ws2[f'C{teach_count}'] = teacher.first
                ws2[f'D{teach_count}'] = teacher.chanceInfected
                teach_count += 1

            assist_count = 2
            ws3 = newfile.create_sheet('Teaching Assistants',2)
            ws3['A1'] = 'Last Name'
            ws3['B1'] = 'First Name'
            ws3['C1'] = 'Infectivity'

            for ta in tas:
                ws3[f'A{assist_count}'] = ta.last
                ws3[f'B{assist_count}'] = ta.first
                ws3[f'C{assist_count}'] = ta.chanceInfected
                assist_count += 1

            basedir = filedir[0:filedir.rindex('\\')]
            filename = f'Result{random.randint(0,100000)}.xlsx'
            newfile.save(os.path.join(basedir, filename))

        else:
            for student in students:
                if f'{student.first} {student.last}' == name:
                    person = student
                    print(f'{name} has a chance of infection of {person.chanceOfDisease}')
                    break

            for teacher in teachers:
                if f'{teacher.first} {teacher.last}' == name:
                    person = teacher
                    print(f'{name} has a chance of infection of {person.chanceInfected}')
                    break

            for ta in tas:
                if f'{ta.first} {ta.last}' == name:
                    person = ta
                    print(f'{name} has a chance of infection of {person.chanceInfected}')
                    break

            if not person:
                print(f'Sorry {name} was not found')


if __name__ == "__main__":
    main()